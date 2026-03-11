"""
File-confirmation report formatters.

Ported from the legacy file_confirmation.py inner functions:
  format_data, format_cptyFile, format_summary_excel, format_fundFile.

The logic is preserved as-is; only the closure dependencies (sDate, runpath,
log_manager, cptyColIdx) are now passed as explicit parameters.
"""

import os
import csv
import datetime
import logging
from typing import Any, Dict, List, Set, Tuple

import openpyxl
import openpyxl.drawing.image
import openpyxl.styles
import openpyxl.utils

logger = logging.getLogger(__name__)


def format_data(
    dataList: list,
    mappingList: list,
    dataHeader: list,
    convertList: dict,
    cptyName: str,
    runpath: str,
    log_manager,
) -> Tuple[list, set]:
    """Map and format raw isql row-lists according to *mappingList* rules.

    Returns ``(formatted_rows, errors)`` where *formatted_rows* is a
    ``list(zip(*columns))`` and *errors* a ``set`` of warning strings.
    """
    formatted: list = []
    err: Set[str] = set()

    for num, i in enumerate(mappingList):
        a = None

        # --- Extract from rawData ---
        if isinstance(i[1], list):
            if i[2] == "str":
                a = list(
                    map(
                        lambda x: list(
                            x[dataHeader.index(y)]
                            for y in i[1]
                            if x[dataHeader.index(y)] != "NULL"
                        )[0],
                        dataList,
                    )
                )
                if len(list(filter(lambda x: x[dataHeader.index(i[1][0])] == "NULL", dataList))) > 0:
                    err.add("Missing " + i[1][0])
            elif i[2] == "nbr":
                a = list(
                    map(lambda x: sum(float(x[dataHeader.index(y)]) for y in i[1]), dataList)
                )
            else:
                print("Undefined list %s" % i[1])
        else:
            a = list(
                map(
                    lambda x: x[dataHeader.index(i[1])] if i[1] in dataHeader else i[1],
                    dataList,
                )
            )

        # --- Format data ---
        if i[2] == "str":
            if isinstance(i[3], list):
                if len(i[3]) == 1:
                    a = list(map(lambda x: x[i[3][0]:], a))
                else:
                    a = list(map(lambda x: x[i[3][0]:i[3][1]], a))
            elif "convert" in i[3]:
                if isinstance(convertList[i[3]], dict):
                    if len(list(filter(lambda x: x not in convertList[i[3]], a))) > 0:
                        err.add("Missing %s mapping" % i[3][7:])
                    a = list(
                        map(
                            lambda x: convertList[i[3]][x] if x in convertList[i[3]].keys() else x,
                            a,
                        )
                    )
                else:
                    if isinstance(convertList[i[3]], list):
                        fName = (
                            convertList[i[3]][0]
                            if os.path.isfile(convertList[i[3]][0])
                            else os.path.join(runpath, convertList[i[3]][0].split(os.sep)[-1])
                        )
                        mapIdx = convertList[i[3]][1]
                        delim = convertList[i[3]][2]
                    else:
                        fName = (
                            convertList[i[3]]
                            if os.path.isfile(convertList[i[3]])
                            else os.path.join(runpath, convertList[i[3]].split(os.sep)[-1])
                        )
                        mapIdx = 1
                        delim = ","

                    dConvert: dict = {}
                    try:
                        with open(fName, "r") as f:
                            reader = csv.reader(f, delimiter=delim)
                            for line in reader:
                                dConvert[line[0]] = line[mapIdx]
                    except IOError as io:
                        err.add("%s mapping file issue" % i[3][7:])
                        log_manager.fastapi_log(
                            "%s::: Fail to open %s mapping file: %s" % (cptyName, i[3][7:], io)
                        )
                    except Exception as e:
                        log_manager.fastapi_log(e)
                        raise e

                    if len(list(filter(lambda x: x not in dConvert, a))) > 0:
                        err.add("Missing %s mapping" % i[3][7:])
                    a = list(map(lambda x: dConvert[x] if x in dConvert else x, a))
            else:
                a = list(map(lambda x: i[3] % x, a))
        elif i[2] == "date":
            a = list(
                map(lambda x: datetime.datetime.strptime("%.8s" % x, "%Y%m%d").strftime(i[3]), a)
            )
        elif i[2] == "nbr":
            a = list(map(lambda x: i[3] % float(x), a))

        formatted.append(a)

    return list(zip(*formatted)), err


def format_cpty_file(
    data: list,
    dConfig: dict,
    sDate: str,
    runpath: str,
    log_manager,
) -> Tuple[str, Any]:
    """Format and write a per-counterparty output file (CSV / TSV / Excel).

    Returns ``(outfile_path, errors)``.
    """
    log_manager.fastapi_log("%s: %d" % (dConfig["cptyName"], len(data) - 1))
    if len(data) <= 1:
        return "", ""

    header = data[0]
    data = data[1:]
    result, err = format_data(
        data,
        dConfig["header"]["headerRow"],
        header,
        {k: val for k, val in dConfig.items() if k.startswith("convert")},
        dConfig["cptyName"],
        runpath,
        log_manager,
    )

    # --- output file name ---
    s: set = set()
    for x in data:
        s.add(x[header.index("CL_TRADE_CURRENCY")][:2])
    sMkt = " ".join(s)
    s.clear()
    for x in data:
        s.add(x[header.index("ACCOUNT")][:-3])
    sFund = " ".join(s)

    out_dir = os.path.join(dConfig["outPath"], dConfig["outFolder"])
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    outprefix = os.path.join(dConfig["outPath"], dConfig["outFolder"], dConfig["outName"]) % {
        "sDate": sDate,
        "sMkt": sMkt,
        "sFund": sFund,
    }
    outfile = outprefix + "." + dConfig["outExt"]
    for i in range(1, 10):
        if os.path.isfile(outfile):
            outfile = (
                outprefix.replace(" ", "").replace("()", "") + "-" + str(i) + "." + dConfig["outExt"]
            )
        else:
            break

    # --- sort ---
    if "sortCol" in dConfig["header"].keys():
        sortIdx = list(zip(*dConfig["header"]["headerRow"]))[0].index(dConfig["header"]["sortCol"])
        sortResult = sorted(result, key=lambda x: x[sortIdx])
    else:
        sortResult = result

    # --- write ---
    if dConfig["outExt"] in ("xls", "xlsx"):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(title="Sheet1", index=0)
        rowConst = 1
        if dConfig["header"]["headerPntFlg"]:
            for j, fld in enumerate(list(zip(*dConfig["header"]["headerRow"]))[0]):
                c = worksheet.cell(column=j + 1, row=1, value=fld)
                c.font = openpyxl.styles.Font(size=11, bold=True, color=openpyxl.styles.colors.BLACK)
                c.fill = openpyxl.styles.PatternFill(
                    fill_type="solid", start_color="aabbcc", end_color="aabbcc"
                )
            rowConst += 1
        for i, row in enumerate(sortResult):
            for j, fld in enumerate(row):
                _ = worksheet.cell(column=j + 1, row=i + rowConst, value=fld)
        workbook.save(outfile)
    else:
        delim = "\t" if dConfig["outExt"] in ("tsv",) else ","
        with open(outfile, "w", newline="") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL, delimiter=delim, quotechar='"')
            if dConfig["header"]["headerPntFlg"]:
                f.write(delim.join(list(zip(*dConfig["header"]["headerRow"]))[0]) + "\n")
            writer.writerows([row for row in sortResult])

    return outfile, err


def format_summary_excel(
    data: list,
    dConfig: dict,
    sDate: str,
    runpath: str,
    log_manager,
) -> Tuple[str, set]:
    """Generate a summary Excel workbook (with subtotals, image, merged header cells).

    Returns ``(outfile_path, errors)``.
    """
    log_manager.fastapi_log("%s: %d - Summary report" % (dConfig["cptyName"], len(data) - 1))
    if len(data) <= 1:
        return "", ""

    ttlConst = "Total"
    headerMapping = dConfig["headerMapping"]
    totalMapping = dConfig["totalMapping"]
    dataMapping = dConfig["dataMapping"]
    outpath = dConfig["outPath"]
    header = data[0]
    data = data[1:]
    result: list = []
    fileHeader: list = []
    summary: list = []
    errs: Set[str] = set()

    imgfile = os.path.join(outpath, "cantor_resize.jpg")

    # --- Record Details ---
    result, err = format_data(
        data,
        dataMapping,
        header,
        {k: val for k, val in dConfig.items() if k.startswith("convert")},
        dConfig["cptyName"],
        runpath,
        log_manager,
    )
    errs.update(err)

    if "sortCol" in dConfig.keys():
        sortIdx = list(zip(*dConfig["dataMapping"]))[0].index(dConfig["sortCol"])
        result = sorted(result, key=lambda x: x[sortIdx])

    fileHeader, err = format_data(
        [data[0]],
        headerMapping,
        header,
        {k: val for k, val in dConfig.items() if k.startswith("convert")},
        dConfig["cptyName"],
        runpath,
        log_manager,
    )
    fileHeader = fileHeader[0]
    errs.update(err)

    # --- subtotal ---
    if "subtotal" in totalMapping:
        filterIdx = list(zip(*dataMapping))[0].index(totalMapping["subtotal"]["filter"])
        filterSet = set(list(zip(*result))[filterIdx])
        if "fixValue" not in totalMapping["subtotal"]:
            totalMapping["subtotal"]["fixValue"] = []
        for i in filterSet | set(x for x in totalMapping["subtotal"]["fixValue"]):
            if i in filterSet:
                filterData = filter(lambda x: x[filterIdx] == i, result)
            aTotal: list = []
            for fld in totalMapping["subtotal"]["header"]:
                cIdx = list(zip(*dataMapping))[0].index(fld)
                countFlg = True if dataMapping[cIdx][1] == "CL_BUY_SELL" else False
                a = (
                    (len(list(filterData)))
                    if countFlg
                    else sum(map(float, list(zip(*filterData))[cIdx]))
                    if i in filterSet
                    else 0
                )
                aTotal.append((a, cIdx))
            summary.append([i, aTotal])
        summary = sorted(summary)
    else:
        log_manager.fastapi_log("No subtotal is required.")

    # --- grand total ---
    if "total" in totalMapping:
        aTotal = []
        for fld in totalMapping["total"]["header"]:
            cIdx = list(zip(*dataMapping))[0].index(fld)
            countFlg = True if dataMapping[cIdx][1] == "CL_BUY_SELL" else False
            a = len(result) if countFlg else sum(map(float, list(zip(*result))[cIdx]))
            aTotal.append((a, cIdx))
        summary.insert(0, [ttlConst, aTotal])
    else:
        log_manager.fastapi_log("No grant total is required.")

    # --- output file name ---
    s: set = set()
    for x in data:
        s.add(x[header.index("CL_TRADE_CURRENCY")][:2])
    sMkt = " ".join(s)
    s.clear()
    for x in data:
        s.add(x[header.index("ACCOUNT")][:-3])
    sFund = " ".join(s)

    out_dir = os.path.join(dConfig["outPath"], dConfig["outFolder"])
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)

    outfile = (
        os.path.join(dConfig["outPath"], dConfig["outFolder"], dConfig["outName"])
        % {"sDate": sDate, "sMkt": sMkt, "sFund": sFund}
        + "."
        + dConfig["outExt"]
    )
    for i in range(1, 30):
        if os.path.isfile(outfile):
            outfile = (
                os.path.join(
                    dConfig["outPath"],
                    dConfig["outFolder"],
                    dConfig["outName"] % {"sDate": sDate, "sMkt": sMkt, "sFund": sFund}
                    + "-"
                    + str(i)
                    + "."
                    + dConfig["outExt"],
                )
            )
        else:
            break

    # --- write workbook ---
    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet(title="Sheet1", index=0)

    cOffset = [7, 1]

    # Summary rows
    sumCell = [cOffset[0] + 1, cOffset[1]]
    for i, row in enumerate(summary):
        tIdx = min(list(zip(*row[1]))[1]) + cOffset[1]
        c = worksheet.cell(row=sumCell[0] + i, column=tIdx, value=row[0])
        c.font = openpyxl.styles.Font(size=10)
        for j, fld in enumerate(row[1]):
            c = worksheet.cell(row=sumCell[0] + i, column=cOffset[1] + fld[1] + 1, value=fld[0])
            c.font = openpyxl.styles.Font(size=10, color="000a90")
        if row[0] == ttlConst:
            for col in range(tIdx, max(list(zip(*row[1]))[1]) + cOffset[1] + 2):
                c = worksheet.cell(column=col, row=sumCell[0] + i)
                c.font = openpyxl.styles.Font(size=10, bold=True, color=openpyxl.styles.colors.BLACK)
                c.fill = openpyxl.styles.PatternFill(
                    fill_type="solid", start_color="aabbcc", end_color="aabbcc"
                )

    cOffset[0] = cOffset[0] + len(summary)

    # Data rows
    for j, fld in enumerate(list(zip(*dataMapping))[0]):
        c = worksheet.cell(row=cOffset[0] + 1, column=j + cOffset[1] + 1, value=fld)
        c.font = openpyxl.styles.Font(size=11, bold=True, color=openpyxl.styles.colors.BLACK)
        c.fill = openpyxl.styles.PatternFill(
            fill_type="solid", start_color="aabbcc", end_color="aabbcc"
        )
    cOffset[0] += 1
    for i, row in enumerate(result):
        for j, fld in enumerate(row):
            _ = worksheet.cell(row=i + cOffset[0] + 1, column=j + cOffset[1] + 1, value=fld)

    # Header section
    for i, fld in enumerate(headerMapping):
        tCell, tMergeOffset, dCell, dMergeOffset = fld[4]
        c = worksheet.cell(column=tCell[1], row=tCell[0], value=fld[0])
        c.font = openpyxl.styles.Font(size=11, bold=True, color=openpyxl.styles.colors.BLACK)
        if tMergeOffset[0] != 0 or tMergeOffset[1] != 0:
            worksheet.merge_cells(
                start_row=tCell[0],
                start_column=tCell[1],
                end_row=tCell[0] + tMergeOffset[0],
                end_column=tCell[1] + tMergeOffset[1],
            )
        sValue = fileHeader[i]
        c = worksheet.cell(column=dCell[1], row=dCell[0], value=sValue)
        if dMergeOffset[0] != 0 or dMergeOffset[1] != 0:
            worksheet.merge_cells(
                start_row=dCell[0],
                start_column=dCell[1],
                end_row=dCell[0] + dMergeOffset[0],
                end_column=dCell[1] + dMergeOffset[1],
            )
        c.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal="left")

    # Column widths
    if "colFitFlg" in dConfig and dConfig["colFitFlg"]:
        column_widths = [5] + list(map(lambda x: 1.5 * max(map(len, x)), list(zip(*result))))
    else:
        column_widths = [5, 20]

    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_width

    # Print settings
    if "printFitFlg" in dConfig and dConfig["printFitFlg"]:
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToPage = True

    # Image (optional – skip if file missing)
    if os.path.isfile(imgfile):
        img = openpyxl.drawing.image.Image(imgfile)
        img.right = True
        worksheet.add_image(img, "B2")

    workbook.template = False
    workbook.save(outfile)

    return outfile, errs


def format_fund_file(
    data: list,
    dConfig: dict,
    cptyColIdx: int,
    grpFxFlg: bool,
    sDate: str,
    runpath: str,
    log_manager,
) -> Tuple[list, set]:
    """Split data by fund / account and delegate to cpty or summary formatter.

    Returns ``(outfiles_list, errors)``.
    """
    log_manager.fastapi_log("%s: %d" % (dConfig["cptyName"], len(data) - 1))
    if len(data) <= 1:
        return "", ""

    result = [data]
    result2: list = []
    for i in dConfig["fileSplit"]["splitBy"]:
        fundIdx = data[0].index(i[0])
        funds = set(list(zip(*data[1:]))[fundIdx])
        fundInc = funds if len(i[1]) == 0 else i[1]
        fundExcl = i[2]
        funds = (set(funds) & set(fundInc)) - set(fundExcl)

        log_manager.fastapi_log("By %s: %s" % (i[0], str(funds)))
        for j in result:
            for k in funds:
                result2.append(
                    list(filter(lambda x: x[fundIdx] == k or x[cptyColIdx] == "COUNTERPARTY_CODE", j))
                )
        result = list(filter(lambda x: len(x) > 1, result2))
        result2 = []

    outfiles: list = []
    errs: Set[str] = set()
    for i in result:
        if grpFxFlg:
            a, b = format_summary_excel(i, dConfig, sDate, runpath, log_manager)
        else:
            a, b = format_cpty_file(i, dConfig, sDate, runpath, log_manager)
        if a != "":
            outfiles.append(a)
        if b != "":
            errs = errs.union(b)

    return outfiles, errs
